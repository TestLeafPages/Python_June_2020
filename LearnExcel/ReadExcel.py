# Step 01:- install openpyxl:- pip install openpyxl

# Read Excel:- import openpyxl

import openpyxl

# Load the workbook (excel)
wb = openpyxl.load_workbook("DemoFile.xlsx")

# returns the sheet names
print(wb.sheetnames)

# return active sheet name
print(wb.active)

# Create the object for sheet
sh = wb['Data_Sheet']

# for row count
print('total rows', sh.max_row)

# for column count
print('total columns', sh.max_column)

# # Way 1:-
# # Read value
# for row in range(2, sh.max_row+1): # for rows
#     for cell in range(1, sh.max_column+1): # for cell
#         v = sh.cell(row, cell).value
#         print(v)

# Way 2:-
# print(sh['C2'].value)
for row in sh['A2' : 'C4']:
    for cell in row:
        print(cell.value)
